# Volatility
# Copyright (C) 2008-2013 Volatility Foundation
# Copyright (C) 2011 Jamie Levy (Gleeda) <jamie.levy@gmail.com>
#
# This file is part of Volatility.
#
# Volatility is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License Version 2 as
# published by the Free Software Foundation.  You may not use, modify or
# distribute this program under any other version of the GNU General
# Public License.
#
# Volatility is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with Volatility.  If not, see <http://www.gnu.org/licenses/>.
#

"""
write to the Free Software
# Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA 02111-1307 USA 
#



Volatile Systems

Edits: March 2013, David Nides edited to include
log2timeline CSV output as option.
Edits: April 2013, Kristinn Gudjonsson edited to include plaso output.
"""


import volatility.plugins.registry.registryapi as registryapi
import volatility.plugins.taskmods as taskmods
import volatility.plugins.registry.shimcache as shimcache
import volatility.plugins.filescan as filescan
import volatility.plugins.sockets as sockets
import volatility.plugins.sockscan as sockscan
import volatility.plugins.modscan as modscan
#import volatility.plugins.iehistory as iehistory
import volatility.plugins.procdump as  procdump
import volatility.plugins.dlldump as dlldump
import volatility.plugins.moddump as moddump
import volatility.plugins.netscan as netscan
import volatility.plugins.evtlogs as evtlogs
import volatility.plugins.userassist as userassist
import volatility.plugins.imageinfo as imageinfo
import volatility.win32.rawreg as rawreg
import volatility.addrspace as addrspace
import volatility.win32.tasks as tasks
import volatility.utils as utils
import volatility.protos as protos
import os, sys
import struct
import volatility.debug as debug
import volatility.obj as obj 
import datetime

try:
    from openpyxl.workbook import Workbook
    from openpyxl.writer.excel import ExcelWriter
    from openpyxl.cell import get_column_letter
    has_openpyxl = True 

except ImportError:
    has_openpyxl = False

try:

  from plaso.events import vol_time
  from plaso.lib import event
  from plaso.lib import storage
  from plaso.lib import preprocess
  import pytz
  import time
except ImportError:
  storage = None

class TimeLiner(dlldump.DLLDump, procdump.ProcExeDump, evtlogs.EvtLogs, userassist.UserAssist):
    """ Creates a timeline from various artifacts in memory """

    def __init__(self, config, *args):  
        evtlogs.EvtLogs.__init__(self, config, *args)
        config.remove_option("SAVE-EVT")
        userassist.UserAssist.__init__(self, config, *args)
        config.remove_option("HIVE-OFFSET")
        config.remove_option("KEY")
        dlldump.DLLDump.__init__(self, config, *args)
        config.remove_option("BASE")
        config.remove_option("REGEX")
        config.remove_option("IGNORE-CASE")
        procdump.ProcExeDump.__init__(self, config, *args)
        config.remove_option("DUMP-DIR")
        config.remove_option("OFFSET")
        config.remove_option("PID")
        config.remove_option("UNSAFE")


        config.add_option('HIVE', short_option = 'H',
                          help = 'Gather Timestamps from a Particular Registry Hive', type = 'str')
        config.add_option('USER', short_option = 'U',
                          help = 'Gather Timestamps from a Particular User\'s Hive(s)', type = 'str')
        config.add_option("REGISTRY", short_option = "R", default = False, action = 'store_true',
                          help = 'Adds registry keys/dates to timeline')

        config.add_option('STORAGE_FILE', short_option = 'S', default = None,
                          type = 'str',
                          help = 'Location of the plaso storage file, so it does not get overwritten.')

    def render_text(self, outfd, data):
        for line in data:
            if line != None:
                outfd.write(line)

    def render_l2tcsv(self, outfd, data):
        for line in data:
            if line != None:
                outfd.write(line)

    def render_body(self, outfd, data):
        for line in data:
            if line != None:
                outfd.write(line)

    def render_plaso(self, outfd, data):
        """Dump events into a Plaso storage file."""
        if not storage:
            raise IOError(
                u"Unable to use plaso as an output, library not available.")

        filename = 'test'
        if not filename:
            raise IOError(
                u"Need to provide a filename to store plaso events.")

        pre_obj = preprocess.PlasoPreprocess()
        if self._config.tz:
            pre_obj.zone = pytz.timezone(self._config.tz)
        else: 


            pre_obj.zone = pytz.utc

        pre_obj.collection_information = {}
        pre_obj.collection_information['output_file'] = filename
        pre_obj.collection_information['file_processed'] = self._config.FILENAME
        pre_obj.collection_information['time_of_run'] = time.time()
        pre_obj.collection_information['method'] = 'volatility collection'
        pre_obj.collection_information['parsers'] = 'Timeliner'

        pathspec = event.EventPathSpec()
        pathspec.type = 'OS'
        pathspec.file_path = filename

        with storage.StorageFile(filename, buffer_size=0, pre_obj=pre_obj) as storage_buffer:
            for item in data:
                item.filename = filename
                item.display_name = u'memory:%s' % filename
                item.pathspec = pathspec
                item.parser = 'Timeliner'
                storage_buffer.AddEntry(item.ToProtoString())

    def render_xlsx(self, outfd, data):
        wb = Workbook(optimized_write = True)
        ws = wb.create_sheet()
        ws.title = 'Timeline Output'
        for line in data:
            coldata = line.split("|")
            ws.append(coldata)
        wb.save(filename = self._config.OUTPUT_FILE)

    def calculate(self):
        if self._config.OUTPUT == "xlsx" and not has_openpyxl:
            debug.error("You must install OpenPyxl for xlsx format:\n\thttps://bitbucket.org/ericgazoni/openpyxl/wiki/Home")
        elif self._config.OUTPUT == "xlsx" and not self._config.OUTPUT_FILE:
            debug.error("You must specify an output *.xlsx file!\n\t(Example: --output-file=OUTPUT.xlsx)")

        if (self._config.HIVE or self._config.USER) and not (self._config.REGISTRY):
            debug.error("You must use -R/--registry in conjuction with -H/--hive and/or -U/--user")

        addr_space = utils.load_as(self._config)
        version = (addr_space.profile.metadata.get('major', 0), 

                   addr_space.profile.metadata.get('minor', 0))

        pids = {}     #dictionary of process IDs/ImageFileName
        offsets = []  #process offsets
        

        im = imageinfo.ImageInfo(self._config).get_image_time(addr_space) 

        body = False
        l2tcsv = False
        plaso = False
        if self._config.OUTPUT == "body":
            body = True
    
        elif self._config.OUTPUT == "l2tcsv":
            l2tcsv = True
        elif self._config.OUTPUT == "plaso":
            plaso = True

        if l2tcsv or plaso:
            if self._config.tz:
                timezone = self._config.tz
            else:
                timezone = 'UTC'

        if l2tcsv:
            line = "date,time,timezone,MACB,source,sourcetype,type,user,host,short,desc,version,filename,inode,notes,format,extra\n"
        elif body:
            line = "0|[END LIVE RESPONSE]|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(im['ImageDatetime'].v())
        else:
            line = "{0}|[END LIVE RESPONSE]\n".format(im['ImageDatetime'])

        if line and not plaso:
            yield line
                

        # Get EPROCESS
        psscan = filescan.PSScan(self._config).calculate()
        for eprocess in psscan:
            if eprocess.obj_offset not in offsets:
                offsets.append(eprocess.obj_offset)

            if l2tcsv:
                unixdatetime = eprocess.CreateTime.v()
                year = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%m/%d/%Y'))
                time = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%H:%M:%S'))
                line = "{0},{1},{6},...B,MEMORY,PROCESS,Created,-,-,,Process: {2}<|>PID: {3}<|>PPID: {4}<|>POffset: 0x{5:08x},,{2},0,-,-,\n".format(
                    year,
                    time,
                    eprocess.ImageFileName,
                    eprocess.UniqueProcessId,
                    eprocess.InheritedFromUniqueProcessId,
                    eprocess.obj_offset,
                    timezone)

                unixdatetime2 = eprocess.ExitTime.v()
                year2 = (datetime.datetime.fromtimestamp(int(unixdatetime2)).strftime('%m/%d/%Y'))
                time2 = (datetime.datetime.fromtimestamp(int(unixdatetime2)).strftime('%H:%M:%S'))
                yield "{0},{1},{6},.A..,MEMORY,PROCESS,Exit,-,-,,Process:{2}<|>PID: {3}<|>PPID: {4}<|>POffset: 0x{5:08x},,{2},0,-,-,\n".format(
                    year2,
                    time2,
                    eprocess.ImageFileName,
                    eprocess.UniqueProcessId,
                    eprocess.InheritedFromUniqueProcessId,
                    eprocess.obj_offset,
                    timezone)
            elif plaso:
                unixdatetime = eprocess.CreateTime.v()
                line = vol_time.EprocessEvent(unixdatetime, eprocess, timezone)
                yield vol_time.EprocessEvent(eprocess.ExitTime.v(), eprocess, timezone, True)
            elif body:
                line = "0|[PROCESS] {2}/PID: {3}/PPID: {4}/POffset: 0x{5:08x}|0|---------------|0|0|0|{0}|{1}|{0}|{0}\n".format(
                    eprocess.CreateTime.v(),
                    eprocess.ExitTime.v(),
                    eprocess.ImageFileName,
                    eprocess.UniqueProcessId,
                    eprocess.InheritedFromUniqueProcessId,
                    eprocess.obj_offset)

            else:
                line = "{0}|{1}|{2}|{3}|{4}|{5}|0x{6:08x}||\n".format(
                    eprocess.CreateTime or '-1',
                    "[PROCESS]",
                    eprocess.ImageFileName,
                    eprocess.UniqueProcessId,
                    eprocess.InheritedFromUniqueProcessId,
                    eprocess.ExitTime or '',
                    eprocess.obj_offset)

            pids[eprocess.UniqueProcessId.v()] = eprocess.ImageFileName
            yield line

        # Get Sockets and Evtlogs XP/2k3 only
        if addr_space.profile.metadata.get('major', 0) == 5:
            socks = sockets.Sockets(self._config).calculate()
            #socks = sockscan.SockScan(self._config).calculate()   # you can use sockscan instead if you uncomment
            for sock in socks:
                la = "{0}:{1}".format(sock.LocalIpAddress, sock.LocalPort)

                if l2tcsv:
                    unixdatetime = sock.CreateTime.v()
                    year = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%m/%d/%Y'))
                    time = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%H:%M:%S'))
                    line = "{0},{1},{7},...B,MEMORY,SOCKET,Created,-,-,,PID: {2}<|>LocalIP: {3}<|>Protocol: {4}({5}),,{2},0,-,-,\n".format(
                        year,
                        time,
                        sock.Pid,
                        la,
                        sock.Protocol,
                        protos.protos.get(sock.Protocol.v(), "-"),
                        sock.obj_offset,
                        timezone)

                elif plaso:
                    unixdatetime = sock.CreateTime.v()
                    line = vol_time.SockEvent(unixdatetime, sock, protos.protos.get(sock.Protocol.v(), "-"))
                elif body:
                    line = "0|[SOCKET] PID: {1}/LocalIP: {2}/Protocol: {3}({4})/POffset: 0x{5:#010x}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                            sock.CreateTime.v(),
                            sock.Pid,
                            la,
                            sock.Protocol,
                            protos.protos.get(sock.Protocol.v(), "-"),
                            sock.obj_offset)

                else:
                    line = "{0}|[SOCKET]|{1}|{2}|Protocol: {3} ({4})|{5:#010x}|||\n".format(
                        sock.CreateTime,
                        sock.Pid,
                        la,
                        sock.Protocol,
                        protos.protos.get(sock.Protocol.v(), "-"),
                        sock.obj_offset)
                yield line

            stuff = evtlogs.EvtLogs.calculate(self)
            for name, buf in stuff:
                for fields in self.parse_evt_info(name, buf, rawtime = True):
                    if l2tcsv:
                        unixdatetime = fields[0]
                        year = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%m/%d/%Y'))
                        time = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%H:%M:%S'))

                        line = "{0},{1},{9},...B,MEMORY,MEM-EVT LOG,Created,-,-,,{2}<|>{3}<|>{4}<|>{5}<|>{6}<|>{7}<|>{8},,-,0,-,-,\n".format(
                            year,
                            time,
                            #Need to replace , for ; so does not throw of delimitor
                            fields[1].replace(",", ";"),
                            fields[2].replace(",", ";"),
                            fields[3].replace(",", ";"),
                            fields[4].replace(",", ";"),
                            fields[5].replace(",", ";"),
                            fields[6].replace(",", ";"),
                            fields[7].replace(",", ";"),
                            timezone)
                    elif plaso:
                        unixdatetime = int(fields[0])
                        line = vol_time.EvtEvent(unixdatetime, fields)
                    elif body:
                        line = '{0} |[EVT LOG]|{1}|{2}|{3}|{4}|{5}|{6}|{7}\n'.format(
                            fields[0], fields[1], fields[2], fields[3], fields[4], fields[5], fields[6], fields[7])
                    else:
                        line = "0|[EVT LOG] {1}/{2}/{3}/{4}/{5}/{6}/{7}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                            fields[0].v(),fields[1], fields[2], fields[3], fields[4], fields[5], fields[6], fields[7])
                    yield line
        else:
            # Vista+
            nets = netscan.Netscan(self._config).calculate()
            for net_object, proto, laddr, lport, raddr, rport, state in nets:
                conn = "{0}:{1} -> {2}:{3}".format(laddr, lport, raddr, rport)
                if l2tcsv:
                    unixdatetime = net_object.CreateTime.v()
                    year = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%m/%d/%Y'))
                    time = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%H:%M:%S'))
                    line = "{0},{1},{7},...B,MEMORY,Network Connection,Created,-,-,,{2}/{3}/{4}/{5}/{6:<#10x},,-,0,-,-,\n".format(
                        year,
                        time,
                        net_object.Owner.UniqueProcessId,
                        conn,
                        proto,
                        state,
                        net_object.obj_offset,
                        timezone)

                elif plaso:
                    unixdatetime = net_object.CreateTime.v()
                    line = vol_time.NetObjectEvent(unixdatetime, net_object, laddr, lport, raddr, rport, proto, state)

                elif body:
                    line = "0|[NETWORK CONNECTION] {1}/{2}/{3}/{4}/{5:<#10x}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                        net_object.CreateTime.v(),
                        net_object.Owner.UniqueProcessId,
                        conn,
                        proto,
                        state,
                        net_object.obj_offset)

                else:
                    line = "{0}|[NETWORK CONNECTION]|{1}|{2}|{3}|{4}|{5:<#10x}||\n".format(
                        str(net_object.CreateTime or "-1"),
                        net_object.Owner.UniqueProcessId,
                        conn,
                        proto,
                        state,
                        net_object.obj_offset)
                yield line

        # Get threads
        threads = modscan.ThrdScan(self._config).calculate()
        for thread in threads:
            image = pids.get(thread.Cid.UniqueProcess.v(), "UNKNOWN")

            if l2tcsv:
                unixdatetime = thread.CreateTime.v()
                year = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%m/%d/%Y'))
                time = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%H:%M:%S'))

                line = "{0},{1},{5},...B,MEMORY,THREAD,Created,-,-,,File: {2}<|>PID: {3}<|>TID: {4},,-,0,-,-,\n".format(
                    year,
                    time,
                    image,
                    thread.Cid.UniqueProcess,
                    thread.Cid.UniqueThread,
                    timezone)

                unixdatetime2 = thread.ExitTime.v()
                year2 = (datetime.datetime.fromtimestamp(int(unixdatetime2)).strftime('%m/%d/%Y'))
                time2 = (datetime.datetime.fromtimestamp(int(unixdatetime2)).strftime('%H:%M:%S'))
                yield "{0},{1},{5},.A..,MEMORY,THREAD,Exit,-,-,,{2}/PID: {3}/TID: {4},,-,0,-,-,\n".format(
                    year2,
                    time2,
                    image,
                    thread.Cid.UniqueProcess,
                    thread.Cid.UniqueThread,
                    timezone)

            elif body:
                line = "0|[THREAD] {2}/PID: {3}/TID: {4}|0|---------------|0|0|0|{0}|{1}|{0}|{0}\n".format(
                    thread.CreateTime.v(),
                    thread.ExitTime.v(),
                    image,
                    thread.Cid.UniqueProcess,
                    thread.Cid.UniqueThread,
                    )
            elif plaso:
                unixdatetime = thread.CreateTime.v()
                yield vol_time.ThreadEvent(unixdatetime, image, thread)
                unixdatetime = thread.ExitTime.v()
                line = vol_time.ThreadEvent(unixdatetime, image, thread, True)

            else:
                line = "{0}|[THREAD]|{1}|{2}|{3}|{4}|||\n".format(
                    thread.CreateTime or '-1',
                    image,
                    thread.Cid.UniqueProcess,
                    thread.Cid.UniqueThread,
                    thread.ExitTime or '',
                    )
            yield line

        # now we get to the PE part.  All PE's are dumped in case you want to inspect them later
        data = moddump.ModDump(self._config).calculate()

        for addr_space, procs, mod_base, mod_name in data:
            space = tasks.find_space(addr_space, procs, mod_base)
            if space != None:
                try:
                    header = procdump.ProcExeDump(self._config).get_nt_header(space, mod_base)
                except ValueError, ve:
                    continue
                try:
                    if l2tcsv:
                        unixdatetime = header.FileHeader.TimeDateStamp.v()
                        year = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%m/%d/%Y'))
                        time = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%H:%M:%S'))
                        line = "{0},{1},{4},...B,MEMORY,MEM-PE Timestamp,Created,-,-,,File: {2}<|>Base: {3:#010x},,-,0,-,-,\n".format(
                            year,
                            time,
                            mod_name,
                            mod_base,
                            timezone)

                    elif plaso:
                        unixdatetime = header.FileHeader.TimeDateStamp.v()
                        line = vol_time.ProcExeDumpEvent(unixdatetime, mod_name, mod_base)

                    elif body:
                        line = "0|[PE Timestamp (module)] {1}/Base: {2:#010x}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                            header.FileHeader.TimeDateStamp.v(),
                            mod_name,
                            mod_base)

                    else:
                        line = "{0}|[PE Timestamp (module)]|{1}||{2:#010x}|||||\n".format(
                            header.FileHeader.TimeDateStamp or '-1',
                            mod_name,
                            mod_base)

                except ValueError, ve:
                    if l2tcsv:
                        pass
                    elif plaso:
                        line = vol_time.ProcExeDumpEvent(0, mod_name, mod_base)

                    elif body:
                        line = "0|[PE Timestamp (module)] {0}/Base: {1:#010x}|0|---------------|0|0|0|0|0|0|0\n".format(
                            mod_name, mod_base)

                    else:
                        line = "-1|[PE Timestamp (module)]|{0}||{1}|||||\n".format(
                            mod_name,
                            mod_base)
                yield line


        # get EPROCESS PE timestamps
        # XXX revert back, now in loop
        for o in offsets:
            self._config.update('OFFSET', o)
            data = self.filter_tasks(procdump.ProcExeDump.calculate(self))
            dllskip = False
            for task in data:
                if task.Peb == None or task.Peb.ImageBaseAddress == None:
                    dllskip = True
                    continue
                try:
                    header = procdump.ProcExeDump(self._config).get_nt_header(task.get_process_address_space(), task.Peb.ImageBaseAddress)
                except ValueError, ve:
                    dllskip = True
                    continue
                try:
                    if l2tcsv:
                        
                        unixdatetime = header.FileHeader.TimeDateStamp.v()
                        year = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%m/%d/%Y'))
                        time = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%H:%M:%S'))
                        line = "{0},{1},{7},...B,MEMORY,MEM-PE Timestamp (exe),Created,-,-,,File: {2}<|>PID: {3}<|>PPID: {4}<|>Command: {5}<|>POffset: 0x{6:08x},,-,0,-,-,\n".format(
                            year,
                            time,
                            task.ImageFileName,
                            task.UniqueProcessId,
                            task.InheritedFromUniqueProcessId,
                            task.Peb.ProcessParameters.CommandLine,
                            o,
                            timezone)

                    elif plaso:
                        unixdatetime = header.FileHeader.TimeDateStamp.v()
                        line = vol_time.PoOffsetEvent(unixdatetime, task, o)

                    elif body:
                        line = "0|[PE Timestamp (exe)] {1}/PID: {2}/PPID: {3}/Command: {4}/POffset: 0x{5:08x}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                            header.FileHeader.TimeDateStamp.v(),
                            task.ImageFileName,
                            task.UniqueProcessId,
                            task.InheritedFromUniqueProcessId,
                            task.Peb.ProcessParameters.CommandLine,
                            o)
                    else:
                        line = "{0}|[PE Timestamp (exe)]|{1}|{2}|{3}|{4}|0x{5:08x}|||\n".format(
                            header.FileHeader.TimeDateStamp or "-1",
                            task.ImageFileName,
                            task.UniqueProcessId,
                            task.InheritedFromUniqueProcessId,
                            task.Peb.ProcessParameters.CommandLine,
                            o)

                except ValueError, ve:
                    if l2tcsv:
                        pass

                    elif plaso:
                        line = vol_time.PoOffsetEvent(0, task, o)

                    elif body:
                        line = "0|[PE Timestamp (exe)] {1}/PID: {2}/PPID: {3}/Command: {4}/POffset: 0x{5:08x}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                            0,
                            task.ImageFileName,
                            task.UniqueProcessId,
                            task.InheritedFromUniqueProcessId,
                            task.Peb.ProcessParameters.CommandLine,
                            o)

                    else:
                        line = "-1|[PE Timestamp (exe)]|{0}|{1}|{2}|{3}|0x{4:08x}|||\n".format(
                            task.ImageFileName,
                            task.UniqueProcessId,
                            task.InheritedFromUniqueProcessId,
                            task.Peb.ProcessParameters.CommandLine,
                            o)
                yield line

            # Get DLL PE timestamps
            if not dllskip:
                dlls = self.filter_tasks(dlldump.DLLDump.calculate(self))
            else:
                dllskip = False
                dlls = []
            for proc, ps_ad, base, basename in dlls:
                if ps_ad.is_valid_address(base):
                    if basename == task.ImageFileName:
                        continue
                    try:
                        header = procdump.ProcExeDump(self._config).get_nt_header(ps_ad, base)
                    except ValueError, ve:
                        continue
                    try:

                        if l2tcsv:
                            unixdatetime = header.FileHeader.TimeDateStamp.v()
                            year = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%m/%d/%Y'))
                            time = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%H:%M:%S'))
                            line = "{0},{1},{8},...B,MEMORY,MEM-PE Timestamp (dll),Created,-,-,,File: {5}<|>Process: {2}<|>PID: {3}<|>PPID: {4}/Process POffset: 0x{6:08x}/DLL Base: 0x{7:8x},,-,0,-,-,\n".format(
                                year,
                                time,
                                task.ImageFileName,
                                task.UniqueProcessId,
                                task.InheritedFromUniqueProcessId,
                                basename,
                                o,
                                base,
                                timezone)

                        elif plaso:
                            unixdatetime = header.FileHeader.TimeDateStamp.v()
                            line = vol_time.DLLEvent(unixdatetime, task, basename, o, base)

                        elif body:
                            line = "0|[PE Timestamp (dll)] {4}/Process: {1}/PID: {2}/PPID: {3}/Process POffset: 0x{5:08x}/DLL Base: 0x{6:8x}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                                header.FileHeader.TimeDateStamp.v(),
                                task.ImageFileName,
                                task.UniqueProcessId,
                                task.InheritedFromUniqueProcessId,
                                basename,
                                o,
                                base)

                        else:
                            line = "{0}|[PE Timestamp (dll)]|{1}|{2}|{3}|{4}|EPROCESS Offset: 0x{5:08x}|DLL Base: 0x{6:8x}||\n".format(
                                header.FileHeader.TimeDateStamp or '-1',
                                task.ImageFileName,
                                task.UniqueProcessId,
                                task.InheritedFromUniqueProcessId,
                                basename,
                                o,
                                base)

                    except ValueError, ve:

                        if l2tcsv:
                            pass

                        elif plaso:
                            line = vol_time.DLLEvent(0, task, basename, o, base)

                        elif body:
                            line = "0|[PE Timestamp (dll)] {4}/Process: {1}/PID: {2}/PPID: {3}/Process POffset: 0x{5:08x}/DLL Base: 0x{6:8x}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                                0,
                                task.ImageFileName,
                                task.UniqueProcessId,
                                task.InheritedFromUniqueProcessId,
                                basename,
                                o,
                                base)
                        else:
                            line = "-1|[PE Timestamp (dll)]|{0}|{1}|{2}|{3}|EPROCESS Offset: 0x{4:08x}|DLL Base: 0x{5:8x}||\n".format(
                                task.ImageFileName,
                                task.UniqueProcessId,
                                task.InheritedFromUniqueProcessId,
                                basename,
                                o,
                                base)
                    yield line

        uastuff = userassist.UserAssist.calculate(self)
        for win7, reg, key in uastuff:
            ts = "{0}".format(key.LastWriteTime)
            for v in rawreg.values(key):
                tp, dat = rawreg.value_data(v)
                subname = v.Name
                if tp == 'REG_BINARY':
                    dat_raw = dat
                    try:
                        subname = subname.encode('rot_13')
                    except UnicodeDecodeError:
                        pass
                    if win7:
                        guid = subname.split("\\")[0]
                        if guid in userassist.folder_guids:
                            subname = subname.replace(guid, userassist.folder_guids[guid])
                    bufferas = addrspace.BufferAddressSpace(self._config, data = dat_raw)
                    uadata = obj.Object("_VOLUSER_ASSIST_TYPES", offset = 0, vm = bufferas)
                    ID = "N/A"
                    count = "N/A"
                    fc = "N/A"
                    tf = "N/A"
                    lw = "N/A"
                    if len(dat_raw) < bufferas.profile.get_obj_size('_VOLUSER_ASSIST_TYPES') or uadata == None:
                        continue
                    else:
                        if hasattr(uadata, "ID"):
                            ID = "{0}".format(uadata.ID)
                        if hasattr(uadata, "Count"):
                            count = "{0}".format(uadata.Count)
                        else:
                            count = "{0}".format(uadata.CountStartingAtFive if uadata.CountStartingAtFive < 5 else uadata.CountStartingAtFive - 5)
                        if hasattr(uadata, "FocusCount"):
                            seconds = (uadata.FocusTime + 500) / 1000.0
                            time = datetime.timedelta(seconds = seconds) if seconds > 0 else uadata.FocusTime
                            fc = "{0}".format(uadata.FocusCount)
                            tf = "{0}".format(time)
                        lw = "{0}".format(uadata.LastUpdated)

                subname = subname.replace("|", "%7c")
                if l2tcsv:
                    unixdatetime = uadata.LastUpdated.v()
                    year = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%m/%d/%Y'))
                    time = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%H:%M:%S'))
                    line = "{0},{1},{8},M...,MEMORY,MEM-USER ASSIST,Last Written,-,-,,{2}/Value: {3}/ID: {4}/Count: {5}/FocusCount: {6}/TimeFocused: {7},,-,0,-,-,\n".format(
                        year,
                        time,
                        reg,
                        subname,
                        ID,
                        count,
                        fc,
                        tf,
                        timezone)

                elif plaso:
                    unixdatetime = uadata.LastUpdated.v()
                    line = vol_time.UserAssistEvent(unixdatetime, reg, subname, ID, count, fc,tf)

                elif body:
                    line = "0|[USER ASSIST] Registry: {1}/Value: {2}/ID: {3}/Count: {4}/FocusCount: {5}/TimeFocused: {6}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                        uadata.LastUpdated.v(), reg, subname, ID, count, fc, tf)
                else:
                    line = "{0}|[USER ASSIST]|{1}|{2}|{3}|{4}|{5}|{6}\n".format(lw, reg, subname, ID, count, fc, tf)
                yield line

        shimdata = shimcache.ShimCache(self._config).calculate()
        for path, lm, lu in shimdata:
            if lu:
                if l2tcsv:
                    #TO DO: Need to yield last update time event
                    unixdatetime = lm.v()
                    year = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%m/%d/%Y'))
                    time = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%H:%M:%S'))
                    line = "{0},{1},{4},...B,MEMORY,MEM-SHIMCACHE,Created,-,-,,{2},,-,0,-,-,\n".format(
                        year,
                        time,
                        path,
                        lu.v(),
                        timezone)

                    unixdatetime2 = lu.v()
                    year2 = (datetime.datetime.fromtimestamp(int(unixdatetime2)).strftime('%m/%d/%Y'))
                    time2 = (datetime.datetime.fromtimestamp(int(unixdatetime2)).strftime('%H:%M:%S'))
                    yield "{0},{1},{4},M...,MEMORY,MEM-SHIMCACHE,Last Written,-,-,,{2},,-,0,-,-,\n".format(
                        year2,
                        time2,
                        path,
                        lu.v(),
                        timezone)
                elif plaso:
                    line = vol_time.ShimCacheEvent(lm.v(), 'created', path)
                    yield vol_time.ShimCacheEvent(lu.v(), 'last_written', path)
                elif body:
                    line = "0|[SHIMCACHE] {1}|0|---------------|0|0|0|{0}|{2}|{0}|{0}\n".format(
                        lm.v(), path, lu.v())
                else:
                    line = "{0}|[SHIMCACHE]|{1}|Last update: {2}\n".format(lm, path, lu)
            else:
                if l2tcsv:
                    pass
                elif plaso:
                    line = vol_time.ShimCacheEvent(lm.v(), 'created', path)
                elif body:
                    line = "0|[SHIMCACHE] {1}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                        lm.v(), path)
                else:
                    line = "{0}|[SHIMCACHE]|{1}|Last update: N/A\n".format(lm, path)
            yield line

        if self._config.REGISTRY:
            regapi = registryapi.RegistryApi(self._config)
            regapi.reset_current()
            regdata = regapi.reg_get_all_keys(self._config.HIVE, self._config.USER, reg = True, rawtime = True)

            for lwtime, reg, item in regdata:
                if l2tcsv:
                    unixdatetime = lwtime.v()
                    year = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%m/%d/%Y'))
                    time = (datetime.datetime.fromtimestamp(int(unixdatetime)).strftime('%H:%M:%S'))
                    line = "{0},{1},{4},M...,MEMORY,MEM-Registry,Last Written,-,-,,{2}<|>{3},,-,0,-,-,\n".format(
                        year,
                        time,
                        reg,
                        item,
                        timezone)

                elif plaso:
                    line = vol_time.RegistryEvent(lwtime.v(), reg, item)
                elif body:
                    line = "0|[REGISTRY] {1}/{2}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                        lwtime.v(), reg, item)

                else:
                    item = item.replace("|", "%7c")
                    line = "{0:<20}|{1}|{2}\n".format(lwtime, reg, item)
                yield line
